import io
import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.schemas.legal import CaseCreate
from app.services.case_candidate_service import CaseCandidateService
from app.services.case_service import CaseService


HEADER_ALIASES = {
    "case_no": {"案号", "案件编号"},
    "plaintiff_name": {"原告", "原告名称"},
    "debtor_name": {"被告", "被告名称", "债务人"},
    "notice_date": {"日期", "通知日期"},
    "remaining_time": {"剩余缴费时间", "缴费期限"},
    "payment_info": {"缴费信息", "缴费金额"},
    "payment_status": {"支付情况", "缴费状态"},
    "tracking_status": {"跟踪情况", "跟进情况"},
}


class CaseImportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.case_service = CaseService(db)

    def parse_xlsx(self, content: bytes, *, group_id: str, tenant_id: str | None = None) -> dict[str, Any]:
        if not content or len(content) > 5 * 1024 * 1024:
            raise ValueError("案件表格不能为空且不得超过 5MB")
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError("无法读取 XLSX 文件") from exc
        sheet = next((item for item in workbook.worksheets if item.max_row >= 1), None)
        if sheet is None:
            raise ValueError("表格中没有可读取的工作表")
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        mapping = self._header_mapping(headers or ())
        missing = [field for field in ("case_no", "debtor_name") if field not in mapping]
        if missing:
            raise ValueError("表格必须包含案号和被告列")

        items: list[dict[str, Any]] = []
        for row_number, values in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            raw = {field: values[index] if index < len(values) else None for field, index in mapping.items()}
            case_no = self.case_service.normalize_case_no(str(raw.get("case_no") or ""))
            debtor = self._clean(raw.get("debtor_name"), 128)
            plaintiff = self._clean(raw.get("plaintiff_name"), 255)
            due_date = self._due_date(raw.get("notice_date"), raw.get("remaining_time"))
            errors: list[str] = []
            if not case_no:
                errors.append("缺少案号")
            if not debtor:
                errors.append("缺少被告")
            existing = self.case_service.find_case_by_case_no(case_no) if case_no else None
            status = "invalid" if errors else "existing" if existing else "create"
            items.append(
                {
                    "row_number": row_number,
                    "status": status,
                    "errors": errors,
                    "existing_case_id": existing.id if existing else None,
                    "case_no": case_no or None,
                    "plaintiff_name": plaintiff,
                    "debtor_name": debtor,
                    "group_id": group_id,
                    "tenant_id": tenant_id,
                    "due_date": due_date.isoformat(),
                    "total_amount": "0.00",
                    "payment_info": self._clean(raw.get("payment_info"), 255),
                    "payment_status": self._clean(raw.get("payment_status"), 64),
                    "tracking_status": self._clean(raw.get("tracking_status"), 500),
                }
            )
        return {
            "sheet_name": sheet.title,
            "total": len(items),
            "creatable": sum(item["status"] == "create" for item in items),
            "existing": sum(item["status"] == "existing" for item in items),
            "invalid": sum(item["status"] == "invalid" for item in items),
            "items": items,
        }

    def confirm(self, preview: dict[str, Any], operator: str) -> dict[str, Any]:
        created_ids: list[int] = []
        for item in preview["items"]:
            if item["status"] != "create":
                continue
            legal_case = self.case_service.create_case(
                CaseCreate(
                    case_no=item["case_no"],
                    debtor_name=item["debtor_name"],
                    plaintiff_name=item.get("plaintiff_name"),
                    group_id=item["group_id"],
                    tenant_id=item.get("tenant_id"),
                    due_date=date.fromisoformat(item["due_date"]),
                    total_amount=Decimal("0.00"),
                    source="spreadsheet_import",
                )
            )
            CaseCandidateService(self.db).resolve_for_existing_case(legal_case, operator)
            created_ids.append(legal_case.id)
        return {**preview, "created": len(created_ids), "created_case_ids": created_ids}

    @staticmethod
    def _header_mapping(headers: tuple[Any, ...]) -> dict[str, int]:
        normalized = {str(value or "").strip(): index for index, value in enumerate(headers)}
        return {
            field: normalized[alias]
            for field, aliases in HEADER_ALIASES.items()
            for alias in aliases
            if alias in normalized
        }

    @staticmethod
    def _clean(value: Any, max_length: int) -> str | None:
        text = str(value or "").strip().rstrip(":：")
        return text[:max_length] or None

    @staticmethod
    def _due_date(notice_date: Any, remaining_time: Any) -> date:
        if isinstance(notice_date, datetime):
            base = notice_date.date()
        elif isinstance(notice_date, date):
            base = notice_date
        else:
            text = str(notice_date or "")
            match = re.search(r"(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})", text)
            base = date(*map(int, match.groups())) if match else date.today()
        days_match = re.search(r"[+＋]?\s*(\d{1,3})\s*天", str(remaining_time or ""))
        return base + timedelta(days=int(days_match.group(1)) if days_match else 30)
